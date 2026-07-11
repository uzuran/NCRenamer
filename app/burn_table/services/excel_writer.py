"""ExcelWriter — appends and updates rows in the burn table workbook."""

from __future__ import annotations

from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from pathlib import Path

    from app.burn_table.models.burn_record import BurnRecord


class TableFullError(Exception):
    """Raised when all 38 data rows (A3:A40) are occupied."""


class ExcelWriter:
    """Writes BurnRecord data into the fixed-layout burn table workbook.

    Both .xls (xlrd + xlutils) and .xlsx (openpyxl) formats are supported.
    All writes are atomic: open → modify → save in a single call.
    """

    DATA_START_ROW = 3
    MAX_ROW = 40

    def __init__(self, sheet_index: int = 0) -> None:
        self._sheet_index = sheet_index

    def append_record(self, path: Path, record: BurnRecord) -> int:
        """Append *record* to the next free row and save.

        Returns the 1-based Excel row number that was written.
        Raises TableFullError if no free row exists.
        """
        if path.suffix.lower() == ".xls":
            return self._append_xls(path, record)
        return self._append_xlsx(path, record)

    def update_record(self, path: Path, row_num: int, record: BurnRecord) -> None:
        """Overwrite an existing row with new data."""
        if not (self.DATA_START_ROW <= row_num <= self.MAX_ROW):
            raise ValueError(
                f"row_num {row_num} is outside data range "
                f"{self.DATA_START_ROW}-{self.MAX_ROW}."
            )
        if path.suffix.lower() == ".xls":
            self._update_xls(path, row_num, record)
        else:
            self._update_xlsx(path, row_num, record)

    def write_record_at_row(self, path: Path, row_num: int, record: BurnRecord) -> None:
        """Write *record* at the specified *row_num* (1-based) and save.

        Used by BurnViewModel when it manages row positioning explicitly (batch
        uploads with a single separator row per batch).  The bounds check is
        delegated to update_record.
        """
        self.update_record(path, row_num, record)

    def ensure_sheet_exists(self, path: Path, sheet_name: str) -> bool:
        """Create the sheet at self._sheet_index if it does not exist yet.

        Returns True if a new sheet was created, False if it already existed.
        """
        if path.suffix.lower() == ".xls":
            return self._ensure_sheet_xls(path, sheet_name)
        return self._ensure_sheet_xlsx(path, sheet_name)

    # ── .xlsx ────────────────────────────────────────────────────────────────

    def _ensure_sheet_xlsx(self, path: Path, sheet_name: str) -> bool:
        import openpyxl

        wb = openpyxl.load_workbook(path)
        if len(wb.worksheets) > self._sheet_index:
            return False  # sheet already exists
        wb.create_sheet(sheet_name)
        wb.save(path)
        return True

    def _append_xlsx(self, path: Path, record: BurnRecord) -> int:
        import openpyxl

        try:
            wb = openpyxl.load_workbook(path)
        except Exception as exc:
            raise ValueError(f"Cannot open workbook '{path}': {exc}") from exc

        ws = wb.worksheets[self._sheet_index]
        next_row = self._find_next_free_xlsx(ws)
        if next_row is None:
            raise TableFullError(
                "The burn table is full (rows A3-A40 are all occupied)."
            )
        self._write_row_xlsx(ws, next_row, record)
        wb.save(path)
        return next_row

    def _update_xlsx(self, path: Path, row_num: int, record: BurnRecord) -> None:
        import openpyxl

        wb = openpyxl.load_workbook(path)
        ws = wb.worksheets[self._sheet_index]
        self._write_row_xlsx(ws, row_num, record)
        wb.save(path)

    def _find_next_free_xlsx(self, ws) -> int | None:
        for row_num in range(self.DATA_START_ROW, self.MAX_ROW + 1):
            if ws.cell(row=row_num, column=2).value is None:
                return row_num
        return None

    def _write_row_xlsx(self, ws, row_num: int, record: BurnRecord) -> None:
        from app.burn_table.services._xlsx_format import (
            make_border,
            make_center_alignment,
        )

        border = make_border()
        center = make_center_alignment()
        for col_idx, value in enumerate(record.to_row(), start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value or None)
            cell.border = border
            cell.alignment = center

    # ── .xls ─────────────────────────────────────────────────────────────────

    def _ensure_sheet_xls(self, path: Path, sheet_name: str) -> bool:
        try:
            import xlrd
            from xlutils.copy import copy as xl_copy
        except ImportError as exc:
            raise ImportError(
                "xlrd and xlutils are required: pip install xlrd==1.2.0 xlutils"
            ) from exc

        rb = xlrd.open_workbook(str(path), formatting_info=True)
        if rb.nsheets > self._sheet_index:
            return False  # sheet already exists
        wb = xl_copy(rb)
        wb.add_sheet(sheet_name)
        wb.save(str(path))
        return True

    def _append_xls(self, path: Path, record: BurnRecord) -> int:
        try:
            import xlrd
            from xlutils.copy import copy as xl_copy
        except ImportError as exc:
            raise ImportError(
                "xlrd and xlutils are required: pip install xlrd==1.2.0 xlutils"
            ) from exc

        try:
            rb = xlrd.open_workbook(str(path), formatting_info=True)
        except Exception as exc:
            raise ValueError(f"Cannot open workbook '{path}': {exc}") from exc

        rs = rb.sheet_by_index(self._sheet_index)
        next_row_idx = self._find_next_free_xls(rs)
        if next_row_idx is None:
            raise TableFullError(
                "The burn table is full (rows A3-A40 are all occupied)."
            )

        wb = xl_copy(rb)
        ws = wb.get_sheet(self._sheet_index)
        self._write_row_xls(ws, next_row_idx, record)
        wb.save(str(path))
        return next_row_idx + 1  # convert 0-based → 1-based Excel row

    def _update_xls(self, path: Path, row_num: int, record: BurnRecord) -> None:
        try:
            import xlrd
            from xlutils.copy import copy as xl_copy
        except ImportError as exc:
            raise ImportError(
                "xlrd and xlutils are required: pip install xlrd==1.2.0 xlutils"
            ) from exc

        rb = xlrd.open_workbook(str(path), formatting_info=True)
        wb = xl_copy(rb)
        ws = wb.get_sheet(self._sheet_index)
        self._write_row_xls(ws, row_num - 1, record)  # 1-based → 0-based
        wb.save(str(path))

    def _find_next_free_xls(self, ws) -> int | None:
        for row_idx in range(self.DATA_START_ROW - 1, self.MAX_ROW):
            if row_idx >= ws.nrows or not str(ws.cell_value(row_idx, 1)).strip():
                return row_idx
        return None

    def _write_row_xls(self, ws, row_idx: int, record: BurnRecord) -> None:
        import xlwt

        # Style must be passed explicitly — calling ws.write() without a style
        # resets the cell to xlwt defaults and destroys the existing borders.
        data_style = xlwt.easyxf(
            "alignment: horiz centre, vert centre;"
            "borders: left thin, right thin, top thin, bottom thin;"
        )
        for col_idx, value in enumerate(record.to_row()):
            ws.write(
                row_idx,
                col_idx,
                value if (value is not None and value != "") else "",
                data_style,
            )

    # ── styled empty separator rows ──────────────────────────────────────────

    def write_empty_row(self, path: Path, row_num: int) -> None:
        """Write a styled empty separator row at *row_num* (1-based Excel row).

        The row is formatted identically to data rows — thin black borders,
        centred alignment, white fill — but contains no values.  Calls with
        *row_num* outside DATA_START_ROW..MAX_ROW are silently ignored.

        Args:
            path:    Path to the .xls or .xlsx burn table workbook.
            row_num: 1-based Excel row to style (e.g. 4 directly after row 3).
        """
        if not (self.DATA_START_ROW <= row_num <= self.MAX_ROW):
            return
        if path.suffix.lower() == ".xls":
            self._write_empty_row_xls(path, row_num)
        else:
            self._write_empty_row_xlsx(path, row_num)

    def _write_empty_row_xlsx(self, path: Path, row_num: int) -> None:
        import openpyxl
        from openpyxl.styles import PatternFill

        from app.burn_table.services._xlsx_format import (
            make_border,
            make_center_alignment,
        )

        wb = openpyxl.load_workbook(path)
        ws = wb.worksheets[self._sheet_index]
        border = make_border()
        center = make_center_alignment()
        fill = PatternFill("solid", fgColor="FFFFFF")
        for col in range(1, 9):  # columns A-H only
            cell = ws.cell(row=row_num, column=col, value=None)
            cell.border = border
            cell.alignment = center
            cell.fill = fill
        wb.save(path)

    def _write_empty_row_xls(self, path: Path, row_num: int) -> None:
        try:
            import xlrd
            import xlwt
            from xlutils.copy import copy as xl_copy
        except ImportError as exc:
            raise ImportError(
                "xlrd and xlutils are required: pip install xlrd==1.2.0 xlutils"
            ) from exc

        rb = xlrd.open_workbook(str(path), formatting_info=True)
        wb = xl_copy(rb)
        ws = wb.get_sheet(self._sheet_index)
        empty_style = xlwt.easyxf(
            "alignment: horiz centre, vert centre;"
            "borders: left thin, right thin, top thin, bottom thin;"
            "pattern: pattern solid, fore_colour white;"
        )
        for col_idx in range(8):  # columns A-H only (0-indexed)
            ws.write(row_num - 1, col_idx, "", empty_style)  # 1-based → 0-based
        wb.save(str(path))

    # ── header migration ─────────────────────────────────────────────────────

    def update_header(self, path: Path) -> None:
        """Rewrite row 1 of an existing file with current headers and formatting.

        All data rows (3-40) are untouched.  Safe to call on old files.
        """
        if path.suffix.lower() == ".xls":
            self._update_header_xls(path)
        else:
            self._update_header_xlsx(path)

    def _update_header_xlsx(self, path: Path) -> None:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill

        from app.burn_table.services._xlsx_format import (
            apply_print_settings,
            make_border,
        )
        from app.burn_table.services.table_factory import (
            _COL_WIDTHS,
            _HEADER_ROW_HEIGHT,
            _ROW1_HEADERS,
        )

        wb = openpyxl.load_workbook(path)
        ws = wb.worksheets[self._sheet_index]

        # Migrate old 10-column format: if F1 is "Čas progr. min", shift data
        # columns G-J left by one (G→F, H→G, I→H, J→I) for all data rows.
        f1 = str(ws.cell(row=1, column=6).value or "").strip().lower()
        if "čas" in f1 and "progr" in f1 and "celkov" not in f1:
            for row_num in range(3, 41):
                for src, dst in ((7, 6), (8, 7), (9, 8), (10, 9)):
                    ws.cell(row=row_num, column=dst).value = ws.cell(
                        row=row_num, column=src
                    ).value

        # Migrate old 9-column format: if C1 is empty (unused note column) and D1
        # is non-empty (sheet_format header), shift data columns D-I → C-H.
        c1 = str(ws.cell(row=1, column=3).value or "").strip()
        d1 = str(ws.cell(row=1, column=4).value or "").strip()
        if not c1 and d1:
            for row_num in range(3, self.MAX_ROW + 1):
                for src, dst in ((4, 3), (5, 4), (6, 5), (7, 6), (8, 7), (9, 8)):
                    ws.cell(row=row_num, column=dst).value = ws.cell(
                        row=row_num, column=src
                    ).value
                ws.cell(row=row_num, column=9).value = None  # clear vacated I

        # Strip columns I and J from all rows — no value, no border, no fill
        no_border = Border()
        no_fill = PatternFill(fill_type=None)
        for row_num in range(1, self.MAX_ROW + 1):
            for col in (9, 10):
                cell = ws.cell(row=row_num, column=col)
                cell.value = None
                cell.border = no_border
                cell.fill = no_fill

        header_font = Font(bold=True, size=9, color="000000")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_fill = PatternFill("solid", fgColor="FFFFFF")
        border = make_border()

        for col_idx, (header, width) in enumerate(
            zip(_ROW1_HEADERS, _COL_WIDTHS, strict=False), start=1
        ):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border
            ws.column_dimensions[cell.column_letter].width = width

        ws.row_dimensions[1].height = _HEADER_ROW_HEIGHT

        # Normalise borders/alignment for ALL data rows A-H so that old files
        # with missing borders on empty rows look consistent.
        from app.burn_table.services._xlsx_format import make_center_alignment

        center = make_center_alignment()
        for row_num in range(self.DATA_START_ROW, self.MAX_ROW + 1):
            for col_num in range(1, 9):
                cell = ws.cell(row=row_num, column=col_num)
                cell.border = border
                cell.alignment = center

        apply_print_settings(ws)
        wb.save(path)

    def _update_header_xls(self, path: Path) -> None:
        try:
            import xlrd
            import xlwt
            from xlutils.copy import copy as xl_copy
        except ImportError as exc:
            raise ImportError(
                "xlrd and xlutils are required: pip install xlrd==1.2.0 xlutils"
            ) from exc

        from app.burn_table.services.table_factory import (
            _COL_WIDTHS,
            _HEADER_ROW_HEIGHT,
            _ROW1_HEADERS,
        )

        rb = xlrd.open_workbook(str(path), formatting_info=True)
        rs = rb.sheet_by_index(self._sheet_index)
        wb = xl_copy(rb)
        ws = wb.get_sheet(self._sheet_index)

        blank_style = xlwt.easyxf("")  # no borders, no fill — truly empty cell
        data_style = xlwt.easyxf(
            "alignment: horiz centre, vert centre;"
            "borders: left thin, right thin, top thin, bottom thin;"
        )

        # ── Step 1: read all data rows into memory (up to 10 old cols) ──────────
        src_row_count = min(self.MAX_ROW, rs.nrows)
        src_col_count = min(10, rs.ncols)
        # data[(row_idx, col_idx)] = cell value (0-based)
        data: dict[tuple[int, int], object] = {}
        for r in range(self.DATA_START_ROW - 1, src_row_count):
            for c in range(src_col_count):
                data[(r, c)] = rs.cell_value(r, c)

        # ── Step 2: apply 10-col migration in memory ─────────────────────────────
        f1_old = (
            str(rs.cell_value(0, 5) if rs.nrows > 0 and rs.ncols > 5 else "")
            .strip()
            .lower()
        )
        if "čas" in f1_old and "progr" in f1_old and "celkov" not in f1_old:
            for r in range(self.DATA_START_ROW - 1, src_row_count):
                for src, dst in ((6, 5), (7, 6), (8, 7), (9, 8)):
                    data[(r, dst)] = data.get((r, src), "")
                data[(r, 9)] = ""

        # ── Step 3: apply 9-col migration in memory ──────────────────────────────
        c1_old = str(
            rs.cell_value(0, 2) if rs.nrows > 0 and rs.ncols > 2 else ""
        ).strip()
        d1_old = str(
            rs.cell_value(0, 3) if rs.nrows > 0 and rs.ncols > 3 else ""
        ).strip()
        if not c1_old and d1_old:
            for r in range(self.DATA_START_ROW - 1, src_row_count):
                for src, dst in ((3, 2), (4, 3), (5, 4), (6, 5), (7, 6), (8, 7)):
                    data[(r, dst)] = data.get((r, src), "")
                data[(r, 8)] = ""

        # ── Step 4: write ALL data rows A-H with consistent borders ──────────────
        # This also fixes old files where A/B had no borders on empty rows.
        for r in range(self.DATA_START_ROW - 1, self.MAX_ROW):
            for c in range(8):
                ws.write(r, c, data.get((r, c), ""), data_style)
            ws.write(r, 8, "", blank_style)  # clear I
            ws.write(r, 9, "", blank_style)  # clear J

        # ── Step 5: write header row ──────────────────────────────────────────────
        header_style = xlwt.easyxf(
            "font: bold True, height 180, colour black;"
            "alignment: wrap True, horiz centre, vert centre;"
            "pattern: pattern solid, fore_colour white;"
            "borders: left thin, right thin, top thin, bottom thin;"
        )
        for col_idx, (header, width) in enumerate(
            zip(_ROW1_HEADERS, _COL_WIDTHS, strict=False)
        ):
            ws.write(0, col_idx, header, header_style)
            ws.col(col_idx).width = width * 256
        ws.write(0, 8, "", blank_style)  # clear I header
        ws.write(0, 9, "", blank_style)  # clear J header

        ws.row(0).height = _HEADER_ROW_HEIGHT * 20
        wb.save(str(path))

    # ── clear all data rows ───────────────────────────────────────────────────

    def clear_all_records(self, path: Path) -> None:
        """Erase all data rows (rows 3-40) from *path*, preserving the header."""
        if path.suffix.lower() == ".xls":
            self._clear_xls(path)
        else:
            self._clear_xlsx(path)

    def _clear_xlsx(self, path: Path) -> None:
        import openpyxl

        wb = openpyxl.load_workbook(path)
        ws = wb.worksheets[self._sheet_index]
        for row_num in range(self.DATA_START_ROW, self.MAX_ROW + 1):
            for col in range(1, 9):
                ws.cell(row=row_num, column=col).value = None
        wb.save(path)

    def _clear_xls(self, path: Path) -> None:
        try:
            import xlrd
            from xlutils.copy import copy as xl_copy
        except ImportError as exc:
            raise ImportError(
                "xlrd and xlutils are required: pip install xlrd==1.2.0 xlutils"
            ) from exc

        rb = xlrd.open_workbook(str(path), formatting_info=True)
        rs = rb.sheet_by_index(self._sheet_index)
        wb = xl_copy(rb)
        ws = wb.get_sheet(self._sheet_index)
        limit = min(self.MAX_ROW, rs.nrows)
        for row_idx in range(self.DATA_START_ROW - 1, limit):
            for col_idx in range(8):
                ws.write(row_idx, col_idx, "")
        wb.save(str(path))

    # ── rewrite (used by delete) ──────────────────────────────────────────────

    def rewrite_all_records(self, path: Path, records: list) -> None:
        """Clear all data rows then write *records* consecutively from DATA_START_ROW.

        Single file open/save — more efficient than clear + N individual writes.
        """
        if path.suffix.lower() == ".xls":
            self._rewrite_all_xls(path, records)
        else:
            self._rewrite_all_xlsx(path, records)

    def _rewrite_all_xlsx(self, path: Path, records: list) -> None:
        import openpyxl

        wb = openpyxl.load_workbook(path)
        ws = wb.worksheets[self._sheet_index]
        # Clear all data rows first
        for row_num in range(self.DATA_START_ROW, self.MAX_ROW + 1):
            for col in range(1, 9):
                ws.cell(row=row_num, column=col).value = None
        # Write records consecutively
        for i, record in enumerate(records):
            self._write_row_xlsx(ws, self.DATA_START_ROW + i, record)
        wb.save(path)

    def _rewrite_all_xls(self, path: Path, records: list) -> None:
        try:
            import xlrd
            from xlutils.copy import copy as xl_copy
        except ImportError as exc:
            raise ImportError(
                "xlrd and xlutils are required: pip install xlrd==1.2.0 xlutils"
            ) from exc

        rb = xlrd.open_workbook(str(path), formatting_info=True)
        rs = rb.sheet_by_index(self._sheet_index)
        wb = xl_copy(rb)
        ws = wb.get_sheet(self._sheet_index)
        # Clear all data rows first
        limit = min(self.MAX_ROW, rs.nrows)
        for row_idx in range(self.DATA_START_ROW - 1, limit):
            for col_idx in range(8):
                ws.write(row_idx, col_idx, "")
        # Write records consecutively
        for i, record in enumerate(records):
            self._write_row_xls(ws, self.DATA_START_ROW - 1 + i, record)
        wb.save(str(path))
