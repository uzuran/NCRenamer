"""ExcelReader - reads the fixed A-J burn table from .xls or .xlsx files."""

from __future__ import annotations

from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from pathlib import Path

from app.burn_table.models.burn_record import BurnRecord


class ExcelReader:
    """Reads BurnRecord rows from the fixed-layout burn table workbook.

    Layout contract:
        - Row 1-2 : header rows  (skipped)
        - Row 3-40: data rows    (up to 38 programs; one blank separator row
          is inserted after each batch, so actual capacity depends on batch sizes)
        - Columns A-J (1-10) map 1-to-1 to BurnRecord fields

    Both .xls (xlrd) and .xlsx (openpyxl) formats are supported.
    """

    DATA_START_ROW = 3
    MAX_ROW = 40
    MAX_DATA_ROWS = 38  # maximum program rows (rows 3-40)

    def __init__(self, sheet_index: int = 0) -> None:
        self._sheet_index = sheet_index

    def find_last_data_row(self, path: Path) -> int | None:
        """Return the 1-based row of the last occupied column-B cell, or None.

        Used by BurnViewModel on load to determine where the next write should
        begin (last_data_row + 2 leaves room for one separator after the batch).
        """
        if path.suffix.lower() == ".xls":
            return self._find_last_data_row_xls(path)
        return self._find_last_data_row_xlsx(path)

    def _find_last_data_row_xlsx(self, path: Path) -> int | None:
        try:
            import openpyxl
        except ImportError as exc:
            raise ImportError("openpyxl is required: pip install openpyxl") from exc
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[self._sheet_index]
        last: int | None = None
        for row_num in range(self.DATA_START_ROW, self.MAX_ROW + 1):
            if ws.cell(row=row_num, column=2).value is not None:
                last = row_num
        wb.close()
        return last

    def _find_last_data_row_xls(self, path: Path) -> int | None:
        try:
            import xlrd
        except ImportError as exc:
            raise ImportError("xlrd is required: pip install xlrd==1.2.0") from exc
        wb = xlrd.open_workbook(str(path))
        ws = wb.sheet_by_index(self._sheet_index)
        last: int | None = None
        limit = min(self.MAX_ROW, ws.nrows)
        for row_idx in range(self.DATA_START_ROW - 1, limit):
            if str(ws.cell_value(row_idx, 1)).strip():
                last = row_idx + 1  # convert 0-based → 1-based
        return last

    def get_existing_programs(self, path: Path) -> list[str]:
        """Return all non-empty program numbers currently stored in *path*."""
        try:
            return [r.program_number for r in self.read_all(path) if r.program_number]
        except Exception:
            return []

    def read_all(self, path: Path) -> list[BurnRecord]:
        """Load every occupied data row from *path* and return as a list."""
        if path.suffix.lower() == ".xls":
            return self._read_xls(path)
        return self._read_xlsx(path)

    def read_all_with_separators(self, path: Path) -> list[BurnRecord | None]:
        """Load data rows and in-batch blank separator rows.

        Returns a mixed list: BurnRecord for data rows, None for blank
        separator rows between batches.  Trailing blank rows (remaining free
        space) are NOT included — they are represented by the gap between
        len(result) and the physical row count.
        """
        if path.suffix.lower() == ".xls":
            return self._read_with_separators_xls(path)
        return self._read_with_separators_xlsx(path)

    # ── .xlsx ────────────────────────────────────────────────────────────────

    def _read_with_separators_xlsx(self, path: Path) -> list[BurnRecord | None]:
        try:
            import openpyxl
        except ImportError as exc:
            raise ImportError("openpyxl is required: pip install openpyxl") from exc

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError(f"Cannot open workbook '{path}': {exc}") from exc

        ws = wb.worksheets[self._sheet_index]
        result: list[BurnRecord | None] = []
        pending_nones: int = 0  # blank rows since the last data row
        for row_num in range(self.DATA_START_ROW, self.MAX_ROW + 1):
            row = [ws.cell(row=row_num, column=col).value for col in range(1, 11)]
            if row[1] is None:
                pending_nones += 1
            else:
                for _ in range(pending_nones):
                    result.append(None)
                pending_nones = 0
                result.append(BurnRecord.from_row(row))
        wb.close()
        # pending_nones at end are free space, not separators — not appended
        return result

    def _read_xlsx(self, path: Path) -> list[BurnRecord]:
        try:
            import openpyxl
        except ImportError as exc:
            raise ImportError("openpyxl is required: pip install openpyxl") from exc

        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError(f"Cannot open workbook '{path}': {exc}") from exc

        ws = wb.worksheets[self._sheet_index]
        records: list[BurnRecord] = []
        for row_num in range(self.DATA_START_ROW, self.MAX_ROW + 1):
            row = [ws.cell(row=row_num, column=col).value for col in range(1, 11)]
            if row[1] is None:  # column B (program_number) is the occupied-row marker
                continue
            records.append(BurnRecord.from_row(row))
        wb.close()
        return records

    # ── .xls ─────────────────────────────────────────────────────────────────

    def _read_with_separators_xls(self, path: Path) -> list[BurnRecord | None]:
        try:
            import xlrd
        except ImportError as exc:
            raise ImportError("xlrd is required: pip install xlrd==1.2.0") from exc

        try:
            wb = xlrd.open_workbook(str(path))
        except Exception as exc:
            raise ValueError(f"Cannot open workbook '{path}': {exc}") from exc

        ws = wb.sheet_by_index(self._sheet_index)
        result: list[BurnRecord | None] = []
        pending_nones: int = 0
        limit = min(self.MAX_ROW, ws.nrows)
        col_count = min(10, ws.ncols)
        for row_idx in range(self.DATA_START_ROW - 1, limit):
            val = ws.cell_value(row_idx, 1)
            if not str(val).strip():
                pending_nones += 1
            else:
                for _ in range(pending_nones):
                    result.append(None)
                pending_nones = 0
                row = [ws.cell_value(row_idx, col) for col in range(col_count)]
                result.append(BurnRecord.from_row(row))
        # pending_nones at end are free space, not separators — not appended
        return result

    def _read_xls(self, path: Path) -> list[BurnRecord]:
        try:
            import xlrd
        except ImportError as exc:
            raise ImportError("xlrd is required: pip install xlrd==1.2.0") from exc

        try:
            wb = xlrd.open_workbook(str(path))
        except Exception as exc:
            raise ValueError(f"Cannot open workbook '{path}': {exc}") from exc

        ws = wb.sheet_by_index(self._sheet_index)
        records: list[BurnRecord] = []
        # xlrd is 0-indexed; cap at ws.nrows so we don't read past the end.
        # Read up to 10 columns but never beyond ws.ncols (new 9-col files have ncols=9).
        limit = min(self.MAX_ROW, ws.nrows)
        col_count = min(10, ws.ncols)
        for row_idx in range(self.DATA_START_ROW - 1, limit):
            val = ws.cell_value(
                row_idx, 1
            )  # column B (program_number) is the occupied-row marker
            if not str(val).strip():
                continue
            row = [ws.cell_value(row_idx, col) for col in range(col_count)]
            records.append(BurnRecord.from_row(row))
        return records
