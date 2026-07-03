"""TableFactory — creates a new, correctly formatted burn table workbook."""

from __future__ import annotations

from pathlib import Path


# Column headers for row 1 — capitalized, single-line, white background
_ROW1_HEADERS = [
    "Datum pálení",               # A
    "Číslo progr.",               # B
    "",                            # C (note / free)
    "Formát Tabule délka × šířka", # D
    "Počet Tabulí Ks",            # E
    "Celkový čas progr.",         # F
    "Vypáleno",                    # G
    "Druh výrobku skupina",       # H
    "Pálil",                       # I
]

# Character widths per column (A–I)
_COL_WIDTHS = [15, 15, 10, 30, 12, 15, 12, 20, 12]

_HEADER_ROW_HEIGHT = 36   # pt — fits two-line text at size 9
_DATA_ROW_HEIGHT   = 18   # pt — compact data rows


class TableFactory:
    """Builds a blank burn table workbook with the correct header layout.

    The resulting file has:
        - Row 1  : main column headers (wrapped text)
        - Row 2  : empty (reserved for sub-headers added manually)
        - Rows 3–36: empty data area
        - Columns A–J fixed widths
    """

    def create(self, path: Path) -> None:
        """Write a new, empty burn table workbook to *path*.

        If a file already exists at *path* it is overwritten.
        Both .xls and .xlsx formats are supported — detected from *path* suffix.
        """
        path.parent.mkdir(parents=True, exist_ok=True)
        if path.suffix.lower() == ".xls":
            self._create_xls(path)
        else:
            self._create_xlsx(path)

    # ── .xlsx ────────────────────────────────────────────────────────────────

    def _create_xlsx(self, path: Path) -> None:
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError as exc:
            raise ImportError("openpyxl is required: pip install openpyxl") from exc

        from ._xlsx_format import apply_print_settings, make_border, make_center_alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pálení"

        header_font  = Font(bold=True, size=9, color="000000")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header_fill  = PatternFill("solid", fgColor="FFFFFF")
        border       = make_border()
        center       = make_center_alignment()

        # ── Row 1: headers ────────────────────────────────────────────────
        for col_idx, (header, width) in enumerate(
            zip(_ROW1_HEADERS, _COL_WIDTHS), start=1
        ):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align
            cell.border    = border
            ws.column_dimensions[cell.column_letter].width = width

        ws.row_dimensions[1].height = _HEADER_ROW_HEIGHT

        # ── Rows 2–36: data area ──────────────────────────────────────────
        for row_num in range(2, 37):
            ws.row_dimensions[row_num].height = _DATA_ROW_HEIGHT
            for col_num in range(1, 10):
                cell            = ws.cell(row=row_num, column=col_num)
                cell.border     = border
                cell.alignment  = center

        apply_print_settings(ws)
        wb.save(path)

    # ── .xls ─────────────────────────────────────────────────────────────────

    def _create_xls(self, path: Path) -> None:
        try:
            import xlwt
        except ImportError as exc:
            raise ImportError("xlwt is required: pip install xlwt") from exc

        wb = xlwt.Workbook(encoding="utf-8")
        ws = wb.add_sheet("Pálení")

        # White background, black bold text, thin borders
        header_style = xlwt.easyxf(
            "font: bold True, height 180, colour black;"   # 180 twips = 9 pt
            "alignment: wrap True, horiz centre, vert centre;"
            "pattern: pattern solid, fore_colour white;"
            "borders: left thin, right thin, top thin, bottom thin;"
        )
        data_style = xlwt.easyxf(
            "alignment: horiz centre, vert centre;"
            "borders: left thin, right thin, top thin, bottom thin;"
        )

        for col_idx, (header, width) in enumerate(zip(_ROW1_HEADERS, _COL_WIDTHS)):
            ws.write(0, col_idx, header.replace("\n", " "), header_style)
            ws.col(col_idx).width = width * 256     # xlwt unit = 1/256 char width

        ws.row(0).height = _HEADER_ROW_HEIGHT * 20  # xlwt unit = 1/20 pt

        # Pre-fill data rows with empty strings + data_style so borders are
        # stored in the file from the start (xlwt does not style empty cells).
        for row_idx in range(1, 36):               # rows 2–36 (0-indexed 1–35)
            ws.row(row_idx).height = _DATA_ROW_HEIGHT * 20
            for col_idx in range(9):
                ws.write(row_idx, col_idx, "", data_style)

        wb.save(str(path))
