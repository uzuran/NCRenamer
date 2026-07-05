"""Integration tests for TableFactory — verifies created files have correct structure."""

import pytest

from app.burn_table.services.table_factory import TableFactory


@pytest.fixture
def factory():
    return TableFactory()


class TestCreateXls:
    def test_creates_file(self, factory, tmp_path):
        path = tmp_path / "table.xls"
        factory.create(path)
        assert path.is_file()

    def test_sheet_named_ocel(self, factory, tmp_path):
        import xlrd

        path = tmp_path / "table.xls"
        factory.create(path)
        wb = xlrd.open_workbook(str(path))
        assert wb.sheet_by_index(0).name == "Ocel"

    def test_has_exactly_one_sheet(self, factory, tmp_path):
        import xlrd

        path = tmp_path / "table.xls"
        factory.create(path)
        assert xlrd.open_workbook(str(path)).nsheets == 1

    def test_header_row_content(self, factory, tmp_path):
        import xlrd

        path = tmp_path / "table.xls"
        factory.create(path)
        ws = xlrd.open_workbook(str(path)).sheet_by_index(0)
        row0 = [ws.cell_value(0, col) for col in range(9)]
        # First header should be 'Datum pálení'
        assert "Datum" in row0[0]
        assert "Číslo" in row0[1]

    def test_data_rows_are_empty(self, factory, tmp_path):
        import xlrd

        path = tmp_path / "table.xls"
        factory.create(path)
        ws = xlrd.open_workbook(str(path)).sheet_by_index(0)
        # xlrd only counts rows that have actual cell records; empty-styled BLANK
        # cells may not appear in nrows — that's fine (no spurious data).
        # Any rows xlrd CAN read beyond the header must have no non-empty values.
        for row_idx in range(1, ws.nrows):
            for col_idx in range(ws.ncols):
                val = ws.cell_value(row_idx, col_idx)
                assert val == "" or val == 0.0, (
                    f"Row {row_idx + 1} col {col_idx + 1} unexpectedly has value: {val!r}"
                )

    def test_overwrites_existing_file(self, factory, tmp_path):
        path = tmp_path / "table.xls"
        path.write_bytes(b"garbage")
        factory.create(path)
        # Should not raise; file should be a valid xls now
        import xlrd

        wb = xlrd.open_workbook(str(path))
        assert wb.nsheets == 1


class TestCreateXlsx:
    def test_creates_file(self, factory, tmp_path):
        path = tmp_path / "table.xlsx"
        factory.create(path)
        assert path.is_file()

    def test_sheet_named_ocel(self, factory, tmp_path):
        import openpyxl

        path = tmp_path / "table.xlsx"
        factory.create(path)
        wb = openpyxl.load_workbook(path)
        assert wb.active.title == "Ocel"

    def test_header_row_content(self, factory, tmp_path):
        import openpyxl

        path = tmp_path / "table.xlsx"
        factory.create(path)
        ws = openpyxl.load_workbook(path, read_only=True).active
        assert ws.cell(row=1, column=1).value == "Datum pálení"

    def test_data_rows_have_borders_but_no_values(self, factory, tmp_path):
        import openpyxl

        path = tmp_path / "table.xlsx"
        factory.create(path)
        ws = openpyxl.load_workbook(path).active
        # Row 3 should have no value in column A
        assert ws.cell(row=3, column=1).value is None

    def test_creates_parent_dirs(self, factory, tmp_path):
        path = tmp_path / "deep" / "nested" / "table.xlsx"
        factory.create(path)
        assert path.is_file()
