"""Integration tests — first launch after a fresh build.

Verifies that when no burn-table Excel file exists:
- load_last_table() auto-creates a blank laser.xlsx at the default location.
- Both Steel (Ocel) and Aluminium (Hliník) sheets exist in the created file.
- Both ViewModels load from the same file after auto-creation.
- No phantom data appears in an empty freshly-created file.
- The created file has the correct header structure.
"""

from __future__ import annotations

from pathlib import Path
from unittest.mock import patch

import pytest

from app.burn_table.services.excel_reader import ExcelReader
from app.burn_table.services.excel_writer import ExcelWriter
from app.burn_table.services.free_slot_detector import FreeSlotDetector
from app.burn_table.viewmodels.performance_recorder import PerformanceRecorder
from app.burn_table.viewmodels.burn_view_model import BurnViewModel
from app.burn_table.viewmodels.print_manager import PrintManager


def _vm(
    sheet_index: int,
    sheet_name: str,
    settings_key: str,
    settings_file: Path | None = None,
) -> BurnViewModel:
    return BurnViewModel(
        reader=ExcelReader(sheet_index=sheet_index),
        writer=ExcelWriter(sheet_index=sheet_index),
        detector=FreeSlotDetector(sheet_index=sheet_index),
        recorder=PerformanceRecorder(),
        print_manager=PrintManager(),
        sheet_name=sheet_name,
        settings_key=settings_key,
        settings_file=settings_file,
    )


class TestFirstLaunchAutoCreate:
    """Simulate a fresh frozen-build first launch with no existing Excel file."""

    def _launch_both_vms(
        self, default_path: Path, settings_file: Path
    ) -> tuple[BurnViewModel, BurnViewModel]:
        """Run load_last_table() on both VMs with no pre-existing file.

        ``settings_file`` is injected at construction time so each VM writes
        the last-opened path to the same per-user file (mirrors App.__init__).
        No module-level patching needed.
        """
        vm_steel = _vm(0, "Ocel", "last_table_path", settings_file=settings_file)
        vm_alu = _vm(1, "Hliník", "last_table_path", settings_file=settings_file)

        # Patch both VMs so _find_existing_table_path returns None (no file)
        # and _default_new_table_path returns our tmp location.
        with (
            patch.object(vm_steel, "_find_existing_table_path", return_value=None),
            patch.object(vm_steel, "_default_new_table_path", return_value=default_path),
            patch.object(vm_alu, "_find_existing_table_path", return_value=None),
            patch.object(vm_alu, "_default_new_table_path", return_value=default_path),
        ):
            vm_steel.load_last_table()
            # After steel creates the file, alu finds it via settings —
            # but with _find_existing_table_path also patched to None for alu,
            # alu would also try to auto-create.  That is idempotent (TableFactory
            # overwrites, then load_table runs).  For the two-VM flow used in
            # App.__init__ the real _find_existing_table_path WOULD find the path
            # saved by steel's _save_settings, so test the real alu path too:
            vm_alu._find_existing_table_path = lambda: (  # type: ignore[method-assign]
                default_path if default_path.is_file() else None
            )
            vm_alu.load_last_table()

        return vm_steel, vm_alu

    def test_excel_file_is_created_on_first_launch(self, tmp_path):
        default = tmp_path / "NCRenamer" / "laser.xlsx"
        settings = tmp_path / "settings.json"
        self._launch_both_vms(default, settings)
        assert default.is_file(), "laser.xlsx must be auto-created on first launch"

    def test_steel_sheet_exists_after_auto_create(self, tmp_path):
        default = tmp_path / "NCRenamer" / "laser.xlsx"
        self._launch_both_vms(default, tmp_path / "settings.json")
        import openpyxl
        wb = openpyxl.load_workbook(default)
        assert "Ocel" in wb.sheetnames

    def test_aluminium_sheet_exists_after_auto_create(self, tmp_path):
        default = tmp_path / "NCRenamer" / "laser.xlsx"
        self._launch_both_vms(default, tmp_path / "settings.json")
        import openpyxl
        wb = openpyxl.load_workbook(default)
        assert "Hliník" in wb.sheetnames

    def test_steel_vm_loads_empty_table(self, tmp_path):
        default = tmp_path / "NCRenamer" / "laser.xlsx"
        vm_steel, _ = self._launch_both_vms(default, tmp_path / "settings.json")
        assert vm_steel.records == [], "freshly-created Steel table must be empty"

    def test_aluminium_vm_loads_empty_table(self, tmp_path):
        default = tmp_path / "NCRenamer" / "laser.xlsx"
        _, vm_alu = self._launch_both_vms(default, tmp_path / "settings.json")
        assert vm_alu.records == [], "freshly-created Aluminium table must be empty"

    def test_both_vms_point_to_same_file(self, tmp_path):
        default = tmp_path / "NCRenamer" / "laser.xlsx"
        vm_steel, vm_alu = self._launch_both_vms(default, tmp_path / "settings.json")
        assert vm_steel.table_path == vm_alu.table_path

    def test_created_file_has_correct_header_in_steel_sheet(self, tmp_path):
        default = tmp_path / "NCRenamer" / "laser.xlsx"
        self._launch_both_vms(default, tmp_path / "settings.json")
        import openpyxl
        wb = openpyxl.load_workbook(default)
        ws = wb["Ocel"]
        # Column A, row 1 must contain "Datum pálení"
        assert ws.cell(row=1, column=1).value == "Datum pálení"

    def test_free_slots_report_38_on_empty_auto_created_table(self, tmp_path):
        default = tmp_path / "NCRenamer" / "laser.xlsx"
        vm_steel, _ = self._launch_both_vms(default, tmp_path / "settings.json")
        assert vm_steel.status.free_rows == 38


class TestDefaultNewTablePath:
    """Unit tests for _default_new_table_path()."""

    def test_returns_none_in_development_mode(self):
        vm = _vm(0, "Ocel", "last_table_path")
        # sys.frozen is not set in dev mode
        result = vm._default_new_table_path()
        assert result is None

    def test_returns_path_in_frozen_mode(self, tmp_path):
        vm = _vm(0, "Ocel", "last_table_path")
        with patch("app.burn_table.viewmodels.burn_view_model.sys") as mock_sys:
            mock_sys.frozen = True
            mock_sys.executable = str(tmp_path / "app.exe")
            with patch.dict("os.environ", {"APPDATA": str(tmp_path)}):
                result = vm._default_new_table_path()
        assert result is not None
        assert result.name == "laser.xlsx"
        assert "NCRenamer" in str(result)

    def test_frozen_path_falls_back_to_home_when_no_appdata(self, tmp_path):
        vm = _vm(0, "Ocel", "last_table_path")
        with patch("app.burn_table.viewmodels.burn_view_model.sys") as mock_sys:
            mock_sys.frozen = True
            mock_sys.executable = str(tmp_path / "app.exe")
            # Remove APPDATA to trigger home-dir fallback
            with patch.dict("os.environ", {}, clear=True):
                with patch("app.burn_table.viewmodels.burn_view_model.Path.home", return_value=tmp_path):
                    result = vm._default_new_table_path()
        assert result is not None
        assert result.suffix == ".xlsx"
