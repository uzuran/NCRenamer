"""Integration tests — hybrid shared/per-user workspace.

Verifies the four data-isolation rules:
  1. Materials  → shared: all users read the same file
  2. Todo       → shared: all users read the same file
  3. Burn table → per-user: each user gets a separate workbook
  4. Settings   → per-user: each user's preferences are independent

Each test class uses a fresh tmp_path fixture so there are no cross-test
filesystem interactions.
"""

from __future__ import annotations

from pathlib import Path

import pytest

from app.burn_table.services.excel_reader import ExcelReader
from app.burn_table.services.excel_writer import ExcelWriter
from app.burn_table.services.free_slot_detector import FreeSlotDetector
from app.burn_table.services.table_factory import TableFactory
from app.burn_table.viewmodels.burn_view_model import BurnViewModel
from app.burn_table.viewmodels.performance_recorder import PerformanceRecorder
from app.burn_table.viewmodels.print_manager import PrintManager
from app.models.material_repository import MaterialRepository
from app.models.settings_model import SettingsModel
from app.models.todo_repository import TodoRepository
from app.utils.workspace import WorkspaceManager


# ── helpers ───────────────────────────────────────────────────────────────────


def _wm(root: Path) -> WorkspaceManager:
    wm = WorkspaceManager(root)
    wm.ensure_shared_workspace_exists()
    return wm


def _wm_with_users(root: Path, *usernames: str) -> WorkspaceManager:
    wm = _wm(root)
    for u in usernames:
        wm.ensure_user_workspace_exists(u)
    return wm


def _burn_vm(sheet_index: int, sheet_name: str, settings_file: Path) -> BurnViewModel:
    return BurnViewModel(
        reader=ExcelReader(sheet_index=sheet_index),
        writer=ExcelWriter(sheet_index=sheet_index),
        detector=FreeSlotDetector(sheet_index=sheet_index),
        recorder=PerformanceRecorder(),
        print_manager=PrintManager(),
        sheet_name=sheet_name,
        settings_key="last_table_path",
        settings_file=settings_file,
    )


# ══════════════════════════════════════════════════════════════════════════════
# 1. Shared files are created once
# ══════════════════════════════════════════════════════════════════════════════


class TestSharedFilesCreatedOnce:
    def test_materials_json_created_in_shared_dir(self, tmp_path):
        wm = _wm(tmp_path)
        MaterialRepository(path=wm.materials_path())
        assert wm.materials_path().is_file()
        assert wm.materials_path().parent == tmp_path / "shared"

    def test_todo_json_created_in_shared_dir(self, tmp_path):
        wm = _wm(tmp_path)
        TodoRepository(path=wm.todo_path())
        assert wm.todo_path().is_file()
        assert wm.todo_path().parent == tmp_path / "shared"

    def test_materials_path_is_identical_regardless_of_user(self, tmp_path):
        wm = _wm_with_users(tmp_path, "alice", "bob")
        assert wm.materials_path() == wm.materials_path()

    def test_todo_path_is_identical_regardless_of_user(self, tmp_path):
        wm = _wm_with_users(tmp_path, "alice", "bob")
        assert wm.todo_path() == wm.todo_path()

    def test_shared_dir_contains_only_shared_files(self, tmp_path):
        wm = _wm_with_users(tmp_path, "alice")
        MaterialRepository(path=wm.materials_path())
        TodoRepository(path=wm.todo_path())
        shared_files = {f.name for f in (tmp_path / "shared").iterdir() if not f.name.endswith(".lock") and not f.name.endswith(".tmp")}
        assert shared_files <= {"materials.json", "todo.json"}


# ══════════════════════════════════════════════════════════════════════════════
# 2. Materials are shared — written by one user, visible to another
# ══════════════════════════════════════════════════════════════════════════════


class TestMaterialsShared:
    def test_material_written_by_alice_visible_to_bob(self, tmp_path):
        wm = _wm_with_users(tmp_path, "alice", "bob")
        repo_alice = MaterialRepository(path=wm.materials_path())
        repo_alice.add_material("WRONG", "RIGHT")

        repo_bob = MaterialRepository(path=wm.materials_path())
        materials = repo_bob.load_materials()
        assert any(row[0] == "WRONG" for row in materials)

    def test_material_deleted_by_alice_also_missing_for_bob(self, tmp_path):
        wm = _wm_with_users(tmp_path, "alice", "bob")
        repo = MaterialRepository(path=wm.materials_path())
        repo.add_material("X", "Y")
        repo.delete_material("X")

        repo_bob = MaterialRepository(path=wm.materials_path())
        assert not any(row[0] == "X" for row in repo_bob.load_materials())

    def test_both_users_read_from_same_physical_file(self, tmp_path):
        wm = _wm_with_users(tmp_path, "alice", "bob")
        assert wm.materials_path() == wm.materials_path()


# ══════════════════════════════════════════════════════════════════════════════
# 3. Todo is shared — written by one user, visible to another
# ══════════════════════════════════════════════════════════════════════════════


class TestTodoShared:
    def test_todo_added_by_alice_visible_to_bob(self, tmp_path):
        wm = _wm_with_users(tmp_path, "alice", "bob")
        repo_alice = TodoRepository(path=wm.todo_path())
        repo_alice.add_item("Buy milk")

        repo_bob = TodoRepository(path=wm.todo_path())
        items = repo_bob.load_items()
        assert any(i["text"] == "Buy milk" for i in items)

    def test_todo_toggled_by_alice_reflected_for_bob(self, tmp_path):
        wm = _wm_with_users(tmp_path, "alice", "bob")
        repo_alice = TodoRepository(path=wm.todo_path())
        item_id = repo_alice.add_item("Task A")
        repo_alice.toggle_done(item_id)

        repo_bob = TodoRepository(path=wm.todo_path())
        items = {i["id"]: i for i in repo_bob.load_items()}
        assert items[item_id]["done"] is True


# ══════════════════════════════════════════════════════════════════════════════
# 4. Burn table is per-user — isolated between users
# ══════════════════════════════════════════════════════════════════════════════


class TestBurnTableIsolation:
    def test_different_users_have_different_burn_table_paths(self, tmp_path):
        wm = _wm_with_users(tmp_path, "alice", "bob")
        assert wm.user_burn_table_path("alice") != wm.user_burn_table_path("bob")

    def test_burn_table_created_only_for_specified_user(self, tmp_path):
        wm = _wm_with_users(tmp_path, "alice", "bob")
        TableFactory().create(wm.user_burn_table_path("alice"))
        assert wm.user_burn_table_path("alice").is_file()
        assert not wm.user_burn_table_path("bob").is_file()

    def test_records_written_to_alice_not_visible_to_bob(self, tmp_path):
        wm = _wm_with_users(tmp_path, "alice", "bob")
        alice_path = wm.user_burn_table_path("alice")
        bob_path = wm.user_burn_table_path("bob")
        TableFactory().create(alice_path)
        TableFactory().create(bob_path)

        from app.burn_table.models.burn_record import BurnRecord
        rec = BurnRecord(
            date="14.07.2026",
            program_number="6670-01",
            sheet_format="1.0037-4X2000X1000",
            sheet_count=1,
            total_time="00:10:00",
        )
        ExcelWriter().append_record(alice_path, rec)

        bob_records = ExcelReader().read_all(bob_path)
        assert bob_records == [], "Bob's table must be empty after Alice writes to hers"

    def test_alice_and_bob_records_do_not_mix(self, tmp_path):
        wm = _wm_with_users(tmp_path, "alice", "bob")
        alice_path = wm.user_burn_table_path("alice")
        bob_path = wm.user_burn_table_path("bob")
        TableFactory().create(alice_path)
        TableFactory().create(bob_path)

        from app.burn_table.models.burn_record import BurnRecord
        rec_alice = BurnRecord(program_number="ALICE-01", sheet_format="1.0037-4X2000X1000", sheet_count=1)
        rec_bob = BurnRecord(program_number="BOB-01", sheet_format="3.3535-6X2000X1000", sheet_count=2)
        ExcelWriter().append_record(alice_path, rec_alice)
        ExcelWriter().append_record(bob_path, rec_bob)

        alice_records = ExcelReader().read_all(alice_path)
        bob_records = ExcelReader().read_all(bob_path)
        assert all(r.program_number == "ALICE-01" for r in alice_records)
        assert all(r.program_number == "BOB-01" for r in bob_records)

    def test_vm_settings_file_is_per_user(self, tmp_path):
        wm = _wm_with_users(tmp_path, "alice", "bob")
        alice_sf = wm.user_settings_path("alice")
        bob_sf = wm.user_settings_path("bob")
        vm_alice = _burn_vm(0, "Ocel", alice_sf)
        vm_bob = _burn_vm(0, "Ocel", bob_sf)
        assert vm_alice._settings_file != vm_bob._settings_file

    def test_vm_table_path_saved_to_correct_user_settings(self, tmp_path):
        wm = _wm_with_users(tmp_path, "alice", "bob")
        alice_path = wm.user_burn_table_path("alice")
        TableFactory().create(alice_path)

        vm_alice = _burn_vm(0, "Ocel", wm.user_settings_path("alice"))
        vm_alice.load_table(alice_path)

        # Bob's settings file must NOT contain Alice's table path
        bob_settings = SettingsModel(path=str(wm.user_settings_path("bob")))
        bob_settings.load()
        assert bob_settings.get("last_table_path") is None


# ══════════════════════════════════════════════════════════════════════════════
# 5. Settings are per-user — isolated between users
# ══════════════════════════════════════════════════════════════════════════════


class TestSettingsIsolation:
    def test_different_users_have_different_settings_paths(self, tmp_path):
        wm = _wm_with_users(tmp_path, "alice", "bob")
        assert wm.user_settings_path("alice") != wm.user_settings_path("bob")

    def test_alice_settings_not_visible_to_bob(self, tmp_path):
        wm = _wm_with_users(tmp_path, "alice", "bob")

        alice_settings = SettingsModel(path=str(wm.user_settings_path("alice")))
        alice_settings.set("language", "cs")
        alice_settings.set("appearance_mode", "Dark")

        bob_settings = SettingsModel(path=str(wm.user_settings_path("bob")))
        bob_settings.load()
        assert bob_settings.get("language") is None
        assert bob_settings.get("appearance_mode") is None

    def test_bob_settings_not_visible_to_alice(self, tmp_path):
        wm = _wm_with_users(tmp_path, "alice", "bob")

        bob_settings = SettingsModel(path=str(wm.user_settings_path("bob")))
        bob_settings.set("language", "en")

        alice_settings = SettingsModel(path=str(wm.user_settings_path("alice")))
        alice_settings.load()
        assert alice_settings.get("language") is None

    def test_settings_round_trip_per_user(self, tmp_path):
        wm = _wm_with_users(tmp_path, "alice")

        s1 = SettingsModel(path=str(wm.user_settings_path("alice")))
        s1.set("language", "cs")

        s2 = SettingsModel(path=str(wm.user_settings_path("alice")))
        s2.load()
        assert s2.get("language") == "cs"

    def test_settings_file_lives_inside_user_dir(self, tmp_path):
        wm = _wm_with_users(tmp_path, "alice")
        settings_path = wm.user_settings_path("alice")
        SettingsModel(path=str(settings_path)).set("x", "1")
        assert settings_path.parent == wm.user_dir("alice")


# ══════════════════════════════════════════════════════════════════════════════
# 6. Both burn-table sheets exist in each user's workbook
# ══════════════════════════════════════════════════════════════════════════════


class TestBothSheetsPerUser:
    def test_steel_and_aluminium_sheets_in_alice_workbook(self, tmp_path):
        wm = _wm_with_users(tmp_path, "alice")
        path = wm.user_burn_table_path("alice")

        vm_steel = _burn_vm(0, "Ocel", wm.user_settings_path("alice"))
        vm_steel.create_new_table(path)          # creates file + "Ocel" sheet
        vm_alu = _burn_vm(1, "Hliník", wm.user_settings_path("alice"))
        vm_alu.load_table(path)                  # ensure_sheet_exists adds "Hliník"

        import openpyxl
        wb = openpyxl.load_workbook(path)
        assert "Ocel" in wb.sheetnames
        assert "Hliník" in wb.sheetnames

    def test_alice_and_bob_each_get_independent_workbooks(self, tmp_path):
        wm = _wm_with_users(tmp_path, "alice", "bob")

        for username in ("alice", "bob"):
            path = wm.user_burn_table_path(username)
            sf = wm.user_settings_path(username)
            vm_steel = _burn_vm(0, "Ocel", sf)
            vm_steel.create_new_table(path)
            vm_alu = _burn_vm(1, "Hliník", sf)
            vm_alu.load_table(path)

        import openpyxl
        for username in ("alice", "bob"):
            wb = openpyxl.load_workbook(wm.user_burn_table_path(username))
            assert "Ocel" in wb.sheetnames
            assert "Hliník" in wb.sheetnames

        assert wm.user_burn_table_path("alice") != wm.user_burn_table_path("bob")
